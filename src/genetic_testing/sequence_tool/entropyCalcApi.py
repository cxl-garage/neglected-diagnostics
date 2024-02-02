import io
import logging
import time

from openpyxl import Workbook, load_workbook

from genetic_testing.sequence_tool.entropyCalcDataClass import *

# from typing import List

# from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class Api:
    @staticmethod
    def calculate(config):
        result = Result()
        state = State(config, result)

        if config.printStream is None:
            config.printStream = print

        calculations = state.factory.get_parser().parse()

        return Api.calculate_common(state, calculations)

    @staticmethod
    def calculate_from_pnns(config, pnns):
        result = Result()
        state = State(config, result)

        result.positions = len(pnns)

        calculations = [
            DoublePositionCalculation(pn.position, pn.a, pn.g, pn.c, pn.u, pn.gap)
            for pn in pnns
        ]

        return Api.calculate_common(state, calculations)

    @staticmethod
    def calculate_common(state, calculations):
        config = state.config
        result = state.result

        print("calculating ...")
        start_time = time.time()

        config.execServcie.map(lambda calc: calc.calculate(), calculations)

        result.time = time.time() - start_time
        config.printStream(f"calculation finished in {result.time} seconds.")

        print("writing the result ...")
        Api.print_result(config, calculations)
        config.printStream("done!")

        return result

    @staticmethod
    def group(config):
        result = Result()
        state = State(config, result)

        result.groups = SequenceFileParser(state).group()
        return result

    @staticmethod
    def pnns(config):
        start_time = time.time()

        result = Result()
        state = State(config, result)

        result.pnns = SequenceFileParser(state).pnns()
        result.time = time.time() - start_time
        result.sequences = result.sequencesTotal
        return result

    @staticmethod
    def create_chart(result):
        if result is None:
            raise ValueError("Result cannot be None")

        wb = load_workbook(
            io.BytesIO(Api.class_loader.get_resource_as_stream("template.xlsx"))
        )
        sheet = wb.active

        result.sort()

        seq_count = len(result)
        for row_index, gs in enumerate(result, start=1):
            row = [gs.get_sequences().size() / seq_count]
            if gs.get_type() == GroupedSequenceType.NORMAL:
                row.insert(0, "")
            else:
                row.append("")
            sheet.append(row)

        wb.save(io.BytesIO())

    @staticmethod
    def create_stratify_result(result, intervals, total_sequences):
        wb = Workbook()
        sheet = wb.active

        # Continue with your implementation here
